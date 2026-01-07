#Function to get max in a list using a linear search
def find_max(data):
    Max = data[0]
    for num in data:
        if num > Max:
            Max = num
    
    return Max


#Emoji converter function
def emoji_converter(message):
    emojiDict = {":)": "😊",
                ":(": "😞",
                ":D": "😃",
                ";)": "😉",
                ":P": "😛",
                ":'(": "😢",
                ":O": "😮",
                ":/": "😕",
                ":|": "😐",
                "<3": "❤️",
                ":*": "😘",
                "XD": "😆",
                "B)": "😎",
                "-_-": "😑",
                "^_^": "😊",
                ">:(": "😠",
                ":3": "😺",
                ":$": "😳",
                ":^)": "🙂"
                }
    res = ""
    for word in message:
        res += emojiDict.get(word, word) + " "
    return res


#Human and Person class definition
class Human:
    def __init__(self, gender):
        self.gender = gender
        
    def breathe(self):
        print(f"{self.gender} Humans breathe")
    def soul(self):
        print(f"{self.gender} Humans have a soul")
    
class Person(Human):
    def __init__(self, name):
        self.name = name

    def talk(self):
        print(f'{self.name} is Talking...')
    
    def walk(self):
        print(f'{self.name} is Walking...')

#Function for the automation of changes to a file in a specific usecase
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename, multi):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    target_col = sheet.max_column + 1
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, target_col-1)
        corrected_price = cell.value * multi
        corrected_price_cell = sheet.cell(row, target_col)
        corrected_price_cell.value = corrected_price

    chartValues = Reference(sheet,
            min_row = 2,
            max_row = sheet.max_row,
            min_col = sheet.max_column,
            max_col = sheet.max_column)

    chart = BarChart()
    
    chart_cell = sheet.cell(2, sheet.max_column+1)

    chart.add_data(chartValues)
    sheet.add_chart(chart, chart_cell.column_letter + '2')

    wb.save(filename)
 
